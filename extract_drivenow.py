from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException

# import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors, Font

import random
import os

from Vehicle import Ride

'''Web Scarping  for www.drivenow.com.au'''
'''
TASKS:
    1. We need to enter pick up date & time and drop off date & time for search window.
    2. Once we choose this and run macro, it should bring up www.drivenow.com.au website
        and enter these dates and time in search window. Pick up location is always Adelaide Airport.
        Then macro automatically click on search.
    3. It will bring up rates for 5 major car rentals, only those rental companies we need to know.

    4. At this stage, we need to scrap data from this webpage back into excel.

'''

# chrome_path = "D:/Softwares/webdriver/chromedriver"
# driver = webdriver.Firefox(
#     executable_path="D:/Softwares/webdriver/geckodriver")


def fetch_data(driver, url):
    driver.get(url)

    # close button banner to check items correctly
    close_banner(driver)

    # set pickup location to Adelaide Airport
    pickup_location = WebDriverWait(driver, default_timeout).until(
        ec.visibility_of_element_located((By.ID, "s0-10-1-1-2-4-4-1-14[0]-3-0-pickupLocation")))
    # pickup_location = driver.find_element(By.ID,
    #                                       "s0-10-1-1-2-4-4-1-14[0]-3-0-pickupLocation")
    pickup_location.clear()
    pickup_location.send_keys("Adelaide Airport (ADL)")

    # select pickup location from suggestion
    if pickup_location.is_selected:
        print("location pickup selected")

        # wait for suggestions to appear
        wait = WebDriverWait(driver, default_timeout)

        # elem = wait.until(ec.visibility_of_element_located(
        #     (By.XPATH,
        #      "//div[@class='autocomplete-suggestions LocationAutocompleteSuggestions LocationAutocompleteSuggestions-s0-10-1-1-2-4-4-1-14[0]-3-0-1-6-pickupLocation X-LocationAutocompleteSuggestions X-LocationAutocompleteSuggestions-Pickup']")))

        elem = wait.until(ec.visibility_of_element_located(
            (By.XPATH, "/html/body/div[3]")))
        # suggestion_elem = driver.find_element_by_xpath("//div[@class='autocomplete-suggestions LocationAutocompleteSuggestions LocationAutocompleteSuggestions-s0-10-1-1-2-4-4-1-14[0]-3-0-1-6-pickupLocation X-LocationAutocompleteSuggestions X-LocationAutocompleteSuggestions-Pickup']").click()

        # move cursor down and hit enter
        pickup_location.send_keys(Keys.DOWN)
        pickup_location.send_keys(Keys.RETURN)
        # pickup_location.click()

        # set PICKUP date
        select_date(driver, pickup_date,
                    "//*[@id='s0-10-1-1-2-4-4-1-14[0]-3-0-pickupDate']")

        # set DROPOFF date
        select_date(driver, dropoff_date,
                    "//*[@id='s0-10-1-1-2-4-4-1-14[0]-3-0-dropOffDate']")

        # dropoff_date_field = WebDriverWait(driver, default_timeout).until(ec.visibility_of_element_located(
        #     (By.XPATH, "//*[@id='s0-10-1-1-2-4-4-1-14[0]-3-0-dropOffDate']")))
        # dropoff_date_field.click()
        # dropoff_date_field.send_keys(DROPOFF_DATE)

    driver.get(url + "/#")
    # submit the form ->
    submit_btn = WebDriverWait(driver, default_timeout).until(ec.visibility_of_element_located(
        (By.XPATH, "//button[@class='btn btn-success btn-lg btn-block X-SearchButton'][@type = 'button']")))
    # submit_btn = driver.find_element_by_class_name("btn btn-success btn-lg btn-block X-SearchButton")
    submit_btn.send_keys(Keys.RETURN)

    # driver.getPageSource()

    WebDriverWait(driver, 10).until(
        ec.presence_of_element_located((By.TAG_NAME, 'body')))
    # driver.implicitly_wait(10)
    # show details in listview

    # list_option = WebDriverWait(driver, 10).until(ec.visibility_of_element_located(
    #     (By.XPATH, "/html/body/div[1]/div[1]/div/div/div[1]/div/div[2]/div/div[1]/div/div/div/div[3]")))
    # list_option.click()

    list_option = WebDriverWait(driver, default_timeout).until(ec.visibility_of_element_located(
        (By.XPATH, "//*[@id='X-Page-Nitro-Content']/div/div/div[1]/div/div[2]/div/div[1]/div/div/div/div[3]")))
    list_option.click()

    # filter supplier for efficiency
    select_supplier(driver)

    # supplier_list = ["Avis", "Budget", "Europecar", "Hertz", "Thrifty"]

    # find Rides
    find_ride(driver)


def select_date(driver, date, xpath):
    date_field = WebDriverWait(driver, default_timeout).until(ec.visibility_of_element_located(
        (By.XPATH, xpath)))
    date_field.click()
    WebDriverWait(driver, default_timeout).until(ec.visibility_of_element_located(
        (By.XPATH, "/html/body/div[1]/div[3]/div[1]/div/div/div[1]/div/div/div/header/button[2]")))
    month_field = driver.find_element(
        By.XPATH, "/html/body/div[1]/div[3]/div[1]/div/div/div[1]/div/div/div/header/button[2]")

    # for setting Pickup date
    # Check for month and select month
    if month_field.text != date.strftime("%B"):
        month_field.click()

        months_field = WebDriverWait(driver, default_timeout).until(ec.visibility_of_element_located(
            (By.XPATH, "/html/body/div[1]/div[3]/div[1]/div/div/div[1]/div/div/div")))
        months = driver.find_elements(
            By.XPATH, "/html/body/div[1]/div[3]/div[1]/div/div/div[1]/div/div/div/button")

        # select current or coming month only
        for i in range(date.month, len(months) + 1):
            single_month = months_field.find_element(
                By.XPATH, "//button[" + str(i) + "]")
            if single_month.text == date.strftime("%B"):
                print("Selected Pickup Month: " + single_month.text)
                single_month.click()
                break
            else:
                print("No any Pickup Month")
    else:
        print("Correct month is selected: " + month_field.text)

    days_field = WebDriverWait(driver, default_timeout).until(ec.visibility_of_element_located(
        (By.XPATH, "/html/body/div[1]/div[3]/div[1]/div/div/div[1]/div/div/div/div")))
    days_list = driver.find_elements(
        By.XPATH, "/html/body/div[1]/div[3]/div[1]/div/div/div[1]/div/div/div/div/button")
    # print("days Field: ", len(days_list))

    for i in range(date.day, len(days_list) + 1):
        try:
            single_day = days_field.find_element(
                By.XPATH, "//button[" + str(i) + "]")
        except StaleElementReferenceException:
            print("StaleElement Exception in selecting day")
        else:
            if single_day.text == str(date.day):
                print("Selected day: " + single_day.text)
                single_day.click()
                if xpath != "//*[@id='s0-10-1-1-2-4-4-1-14[0]-3-0-dropOffDate']":
                    WebDriverWait(driver, default_timeout).until(ec.visibility_of_element_located(
                        (By.XPATH, "/html/body/div[1]/div[3]/div[1]/div/div/footer/a"))).click()
                break


# supplier_list = ['Avis', 'Budget', 'Euro']
supplier_list = []

exception_count = 0


def select_supplier(driver):
    table_field = "//*[@id='X-Page-Nitro-Content']/div/div/div[1]/div/div[2]/div/div[2]/div/div/div/div[2]/span[4]/div/div[1]/table/"

    alamo_field = "img[@src='/webdata/images/supplier/logo/alamo.gif']"
    avis_field = "img[@src='/webdata/images/supplier/logo/lrg/avis.gif']"
    budegt_field = "img[@src='/webdata/images/supplier/logo/budget-140-40.jpg']"
    dollar_field = "img[@src='/webdata/images/supplier/logo/lrg/dollar.gif']"
    enterprise_field = "img[@src='/webdata/images/supplier/logo/enterprise.gif']"
    europcar_field = "img[@src='/webdata/images/supplier/logo/lrg/europcar.gif']"
    firefly_field = "img[@src='/webdata/images/supplier/logo/lrg/firefly.gif']"
    hertz_field = "img[@src='/webdata/images/supplier/logo/hertz.gif']"
    keddy_field = "img[@src='/webdata/images/supplier/logo/lrg/keddy.gif']"
    thrifty_field = "img[@src='/webdata/images/supplier/logo/lrg/thrifty.gif']"

    suppliers = [alamo_field, avis_field, budegt_field, dollar_field, enterprise_field,
                 europcar_field, firefly_field, hertz_field, keddy_field, thrifty_field]

    supplier_list = WebDriverWait(driver, 10).until(ec.visibility_of_element_located(
        (By.XPATH, "//div[@class='X-DnSelectionMenu dn-selection-menu supplier X-supplier']")))
    supplier_list.click()

    # check for budget
    # check_supplier(
    #     driver, "//span/img[@src='/webdata/images/supplier/logo/budget-140-40.jpg']")

    for i in selected_supplier:
        check_supplier(driver, "//span/" + suppliers[i])

    # check avis if present
    # check_supplier(
    #     driver, "//span/" + avis_field)

    # check budget if present
    # check_supplier(
    #     driver, "//span/" + budegt_field)

    # # check europcar if present
    # check_supplier(
    #     driver, "//td[2]/span/" + europcar_field)

    # # check and mark hertz if present
    # check_supplier(
    #     driver, "//td[2]/span/" + hertz_field)

    # # check if thrifty is present and mark if present
    # check_supplier(driver, "//td[2]/span/" + thrifty_field)

    total_supplier = 5

    if exception_count == total_supplier:
        # insert all suppliers name as error in filter selection
        table_path = "//*[@id='X-Page-Nitro-Content']/div/div/div[1]/div/div[2]/div/div[2]/div/div/div/div[2]/span[4]/div/div[1]/table"
        suppliers_table = driver.find_elements(By.XPATH, table_path)

        # skip all top first row i.e alll supplier option
        for i in range(2, len(suppliers_table) + 1):
            # find supplier img and parse supplier name
            try:
                supplier_img = suppliers_table.find_element(
                    By.XPATH, table_path + '/tr[' + str(i) + ']/span/img')
                supplier_name = parse_supplier_name_from_img(supplier_img)
                supplier_list.append(supplier_name)
                print("Supplier List: " + supplier_list)
            except (NoSuchElementException, AttributeError):
                print("Select Supplier Table Img Error:" + str(i))

    # too slow while running in loops
    # for i in range(1, 10):

    #     try:
    #         # check budget if present
    #         check_supplier(
    #             driver, table + "tr[" + str(i) + "]/td[2]/span/img[@src='/webdata/images/supplier/logo/budget-140-40.jpg']")
    #     except TimeoutException:
    #         print("Timed Out for Budget: " + str(i))

    #     try:
    #         # check europcar if present
    #         check_supplier(
    #             driver, table + "tr[" + str(i) + "]/td[2]/span/img[@src='/webdata/images/supplier/logo/lrg/europcar.gif']")

    #     except TimeoutException:
    #         print("Timed Out for Europcar: " + str(i))

    #     try:
    #         # check and mark hertz if present
    #         check_supplier(
    #             driver, table + "tr[" + str(i) + "]/td[2]/span/img[@src='/webdata/images/supplier/logo/hertz.gif']")
    #     except TimeoutException:
    #         print("Timed Out for Hertz: " + str(i))

    #     try:
    #         # check if thrifty is present and mark if present
    #         check_supplier(driver, table + "tr[" + str(
    #             i) + "]/td[2]/span/img[@src='/webdata/images/supplier/logo/lrg/thrifty.gif']")
    #     except TimeoutException:
    #         print("Timed Out for Thrifty: " + str(i))

    # thrifty = WebDriverWait(driver, 10).until(ec.visibility_of_element_located(
    #     (By.XPATH, table + "tr[9]/td[2]/span/img[@src='/webdata/images/supplier/logo/lrg/thrifty.gif']")))
    # if thrifty:
    #     driver.find_element(By.XPATH, table+ "tr[9]").click()


# to filter supplier
def check_supplier(driver, xpath):
    global exception_count
    # check item if present
    try:
        item = WebDriverWait(driver, 1).until(ec.visibility_of_element_located(
            (By.XPATH, xpath)))
        # raise TimeoutException
    except TimeoutException:
        srcs = xpath.split("/")
        # split img name and format and select image name
        supplier_name = srcs[-1].split(".")[0]
        # split for budget-140-40
        supplier_name = supplier_name.split('-')[0]
        print("Time out for: " + supplier_name)
        exception_count += 1
    else:
        if item:
            item.click()
            suppplier_name = parse_supplier_name_from_img(item)
            supplier_list.append(suppplier_name)


def close_banner(driver):
    WebDriverWait(driver, 10).until(ec.visibility_of_element_located(
        (By.XPATH, "/html/body/div[1]/div[2]/div/div[1]"))).click()


def find_ride(driver):
    # rides = WebDriverWait(driver, 10).until(ec.visibility_of_element_located(
    #     (By.XPATH, "/html/body/div[1]/div[1]/div/div/div[1]/div/div[4]/div")))

    rides_field = WebDriverWait(driver, 10).until(ec.visibility_of_element_located(
        (By.XPATH, "//*[@id='X-Page-Nitro-Content']/div/div/div[1]/div/div[4]/div")))

    rides_list = rides_field.find_elements(
        By.XPATH, "//div/div[@class='car-result']")

    currency = rides_field.find_element(
        By.XPATH, "//*[@id='X-Page-Nitro-Content']/div/div/div[1]/div/div[4]/div/div/div/div[1]/div/div/table[1]/tr/td[3]/div[1]/span/span/span[1]/span")
    # print(rides_field.text)

    # all_rides = driver.find_elements(By.XPATH, "//*[@id='X-Page-Nitro-Content']/div/div/div[1]/div/div[4]/div")
    print("Total Rides: ", len(rides_list))

    data_path = "//*[@id='X-Page-Nitro-Content']/div/div/div[1]/div/div[4]/div/div/div/div["

    price_heading = 'price (' + currency.text + ')'
    # create pandas file
    # df = pd.DataFrame(columns=['Name', 'Supplier', 'Type', price_heading])

    rides = []  # list all ride objects
    for i in range(1, len(rides_list) + 1):

        try:
            price = driver.find_element(
                By.XPATH, data_path + str(i) + "]/div/div/table[1]/tr/td[3]/div[1]/span/span/span[2]")
            print("Price: " + str(i) + " : " + price.text)

        except NoSuchElementException:
            print("Element not found Price:" + str(i))

        try:
            ride_name = rides_field.find_element(
                By.XPATH, data_path + str(i) + "]/div/div/table[1]/tr/td[2]/div[1]/span")
        except NoSuchElementException:
            print("Element not found Name:" + str(i))

        try:
            supplier_img = rides_field.find_element(
                By.XPATH, data_path + str(i) + "]/div/div/table[2]/tr/td[1]/div/img")
            supplier_name = parse_supplier_name_from_img(supplier_img)
            # src = supplier_img.get_attribute('src')
            # srcs = src.split("/")
            # # split img name and format and select image name
            # supplier_name = srcs[-1].split(".")[0]
            # # split for budget-140-40
            # supplier_name = supplier_name.split('-')[0]
            # supplier_name = supplier_name.capitalize()
        except NoSuchElementException:
            print("Element not found Supplier Img:" + str(i))

        try:
            vehicle_type = rides_field.find_element(
                By.XPATH, data_path + str(i) + "]/div/div/table[1]/tr/td[2]/div[3]")
        except NoSuchElementException:
            print("Element not found Vehicle Type:" + str(i))

        else:
            print(ride_name.text + " supplier: " +
                  supplier_name + " vehicle Type: " + vehicle_type.text)

            ride = Ride(ride_name.text,
                        supplier_name, vehicle_type=vehicle_type.text, price=price.text)

            rides.append(ride)

            # data = {'Name': ride_name.text, 'Supplier': supplier_name, 'Type':vehicle_type.text, price_heading: price.text}
            # new_data = pd.Series(data)
            # df = df.append(pd.Series(data), ignore_index=True)

    # df.to_csv('data.csv')
    # print(rides)
    write_to_excel(rides, supplier_list, price_heading)


def parse_supplier_name_from_img(img_field):
    src = img_field.get_attribute('src')
    srcs = src.split("/")
    # split img name and format and select image name
    supplier_name = srcs[-1].split(".")[0]
    # split for budget-140-40
    supplier_name = supplier_name.split('-')[0]
    supplier_name = supplier_name.capitalize()
    return supplier_name


def write_to_excel(rides, supplier_list, price_heading):

    workbook = Workbook()
    worksheet = workbook.active

    redFill = PatternFill(start_color='FF0000',
                          end_color='FF0000',
                          fill_type='solid')
    blueFill = PatternFill(start_color='00B0F0',
                           end_color='00B0F0',
                           fill_type='solid')
    purpleFill = PatternFill(start_color='213764',
                             end_color='213764',
                             fill_type='solid')

    yellowFill = PatternFill(start_color='FFFF00',
                             end_color='FFFF00',
                             fill_type='solid')
    greenFill = PatternFill(start_color='009B4E',
                            end_color='009B4E',
                            fill_type='solid')

    color_list = [redFill, blueFill, purpleFill]

    # Add a bold format to use to highlight cells.

    worksheet['A1'] = "PICKUP DATE"
    worksheet['A1'].font = Font(bold=True)
    worksheet['A2'] = "PICKUP TIME"
    worksheet['A2'].font = Font(bold=True)
    worksheet['A4'] = "DROPOFF DATE"
    worksheet['A4'].font = Font(bold=True)
    worksheet['A5'] = "DROPOFF TIME"
    worksheet['A5'].font = Font(bold=True)

    worksheet['C1'] = PICKUP_DATE
    worksheet['C1'].font = Font(bold=True)
    worksheet['C4'] = DROPOFF_DATE
    worksheet['C4'].font = Font(bold=True)

    row = 9
    col = 3
    init_col = 3
    color_choice = color_list[0]
    prev_choice = color_list[1]
    for supplier in supplier_list:
        worksheet.cell(row, col, value=supplier)
        color_choice = random.choice(color_list)
        while color_choice == prev_choice:
            color_choice = random.choice(color_list)
        prev_choice = color_choice
        worksheet.cell(row, col).fill = color_choice
        worksheet.cell(row, col).font = Font(bold=True, color=colors.WHITE)
        col += 1

    # merge supplier name cell
    worksheet.merge_cells(
        start_row=row - 1, start_column=init_col, end_row=row - 1, end_column=col - 1)
    # worksheet.merge_range(row - 1, init_col, row - 1,
    #                       col - 1, 'Supplier', bold)

    # worksheet.write('A9', 'S.No.', bold)
    worksheet['A8'] = 'Name'
    worksheet['A8'].font = Font(bold=True)

    worksheet['B8'] = 'Type'
    worksheet['B8'].font = Font(bold=True)

    worksheet['C8'] = 'Supplier ' + price_heading
    worksheet['C8'].font = Font(bold=True)

    # prev_vehicle_type = ""
    # excel_list = []
    excel_list = {}
    row, i = 10, 1
    for ride in rides:
        col = 1

        # rows = [worksheet.rows]
        is_added = False
        line = 10
        for item in worksheet.rows:

            if worksheet.cell(row=line, column=1).value == ride.get_name() and worksheet.cell(line, 2).value == ride.get_vehicle_type():
                print("Added: in " + str(line) + ": " + ride.get_name())
                i = 3
                for supplier in supplier_list:
                    if ride.get_supplier() == supplier:
                        worksheet.cell(line, i, value=float(ride.get_price()))
                        # set added to true
                        print("found supplier: " + supplier)
                        is_added = True
                        break
                    i += 1

            line += 1
        if not is_added:
            worksheet.cell(row, col, ride.get_name())

            col += 1
            worksheet.cell(row, col, ride.get_vehicle_type())
            col += 1
            # worksheet.cell(row, col, ride.get_supplier())

            write_supplier_price(worksheet, row, ride)
            col += 4
            # worksheet.cell(row, col, ride.get_price())
            price_col = col
            col += 1

            excel_list.update({row: ride})
            # print("excel dict: " , len(excel_list))

            row += 1
            i += 1

    # row, i = 10, 0
    # for ride in rides:

    # worksheet.cell(8, price_col, price_heading)
    dir = "data"
    filename = 'ride_list_of_' + PICKUP_DATE.lower() + ' to ' + DROPOFF_DATE.lower() + '.xlsx'
    if not os.path.exists(dir):
        os.makedirs(dir)
    distinct_filename = get_nonexistant_path(dir + '/' + filename)
    workbook.save(distinct_filename)
    print("Workbook Saved: " + distinct_filename)

def get_nonexistant_path(fname_path):
    """
    Get the path to a filename which does not exist by incrementing path.

    Example
    --------
    >>> get_nonexistant_path('rides.xlsx)
    'rides1.xlsx'
    """
    if not os.path.exists(fname_path):
        return fname_path
    filename, file_extension = os.path.splitext(fname_path)
    i = 1
    new_fname = "{}-{}{}".format(filename, i, file_extension)
    while os.path.exists(new_fname):
        i += 1
        new_fname = "{}-{}{}".format(filename, i, file_extension)
    return new_fname

def write_supplier_price(worksheet, row, ride):
    i = 3
    for supplier in supplier_list:
        if ride.get_supplier() == supplier:
            worksheet.cell(row, i, value=float(ride.get_price()))
            # set added to true
            # print("found supplier: "+ supplier)
            break
        i += 1


'''
    INSERT FIELDS...
'''

''' Enter Pickup date and time here, format-> Thu 22 Aug 2019 '''


current_date = datetime.now()

# assign pickup date of  tomorrow, set from_day = 1
# and for next days in increasing order 2, 3, 4 respectively.
from_day = 3
# assign total days for vehicle booking
total_day_for_booking = 2

pickup_date = current_date + timedelta(days=from_day)
PICKUP_DATE = pickup_date.strftime("%a %d %b %Y")
print("PICKUP DATE: " + PICKUP_DATE)
PICKUP_TIME = ""

''' Enter Dropoff date and time here : format -> Fri 23 Aug 2019 '''

# add total days for booking in pickup days
dropoff_date = pickup_date + timedelta(days=total_day_for_booking)
dropoff_date = pickup_date + timedelta(days=60)
DROPOFF_DATE = dropoff_date.strftime("%a %d %b %Y")
print("DROPOFF DATE: " + DROPOFF_DATE)
DROPOFF_TIME = ""

'''
[alamo_field = 0, avis_field = 1, 
budegt_field = 2, dollar_field = 3, enterprise_field = 4,
 europcar_field = 5, firefly_field = 6, 
hertz_field = 7, keddy_field = 8, thrifty_field = 9]

'''
# enter the index of supplier you want to select in serial order in selected supplier list below
# select all
# selected_supplier = [x for x in range(0, 10)]

# select avis, budget and dollar
selected_supplier = [1, 2, 3, 0, 4, 5]



# timeout for webdriverWait
default_timeout = 10

if __name__ == '__main__':
    # For launching in Chrome
    # make chrome browser stay open after executing code
    from selenium.webdriver.chrome.options import Options
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_experimental_option("detach", True)
    chrome_path = "chromedriver"
    driver = webdriver.Chrome(
        executable_path=chrome_path, chrome_options=chrome_options)

    # FOR launching firefox,comment above chrome code &  uncomment below lines
    # driver = webdriver.Firefox(
    #     executable_path="D:/Softwares/webdriver/geckodriver")
    url = "https://www.drivenow.com.au"

    try:
        fetch_data(driver, url)
    finally:
        # driver.quit()
        pass